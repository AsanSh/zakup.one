"""
Утилиты для генерации счетов и документов
"""
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from django.core.files.base import ContentFile
from django.utils import timezone


def generate_invoice_pdf(order):
    """Генерация PDF счета на оплату"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    story = []
    
    # Заголовок
    story.append(Paragraph('СЧЕТ НА ОПЛАТУ', title_style))
    story.append(Spacer(1, 10*mm))
    
    # Информация о счете
    invoice_data = [
        ['Номер счета:', order.invoice_number or f'INV-{order.order_number}'],
        ['Дата:', datetime.now().strftime('%d.%m.%Y')],
        ['Номер заказа:', order.order_number or str(order.id)],
    ]
    
    invoice_table = Table(invoice_data, colWidths=[60*mm, 120*mm])
    invoice_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(invoice_table)
    story.append(Spacer(1, 10*mm))
    
    # Информация о компании
    if order.company_name:
        company_data = [
            ['Плательщик:', order.company_name],
            ['ИНН:', order.company_inn or ''],
            ['Банк:', order.company_bank or ''],
            ['Расчетный счет:', order.company_account or ''],
            ['Юридический адрес:', order.company_legal_address or ''],
        ]
        
        company_table = Table(company_data, colWidths=[60*mm, 120*mm])
        company_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(company_table)
        story.append(Spacer(1, 10*mm))
    
    # Таблица товаров
    items_data = [['№', 'Наименование', 'Кол-во', 'Ед.', 'Цена', 'Сумма']]
    
    for idx, item in enumerate(order.items.all(), 1):
        items_data.append([
            str(idx),
            item.product.name,
            str(item.quantity),
            item.product.unit,
            f"{float(item.price):,.2f}",
            f"{float(item.total_price):,.2f}"
        ])
    
    # Итого
    items_data.append(['', '', '', '', 'ИТОГО:', f"{float(order.total_amount):,.2f}"])
    
    items_table = Table(items_data, colWidths=[10*mm, 70*mm, 20*mm, 15*mm, 30*mm, 30*mm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('ALIGN', (4, 0), (5, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('FONTNAME', (4, -1), (5, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (4, -1), (5, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 10*mm))
    
    # Подпись
    story.append(Paragraph('Всего к оплате: <b>{:,.2f} сом</b>'.format(float(order.total_amount)), styles['Normal']))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_invoice_excel(order):
    """Генерация Excel файла со списком товаров"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товарная спецификация"
    
    # Заголовок
    ws['A1'] = 'ТОВАРНАЯ СПЕЦИФИКАЦИЯ'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Информация о заказе
    ws['A3'] = 'Номер заказа:'
    ws['B3'] = order.order_number or str(order.id)
    ws['A4'] = 'Дата:'
    ws['B4'] = datetime.now().strftime('%d.%m.%Y')
    
    if order.company_name:
        ws['A5'] = 'Компания:'
        ws['B5'] = order.company_name
    
    # Заголовки таблицы
    headers = ['№', 'Наименование', 'Артикул', 'Количество', 'Ед.', 'Цена', 'Сумма']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = openpyxl.styles.PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    
    # Данные товаров
    row = 8
    for idx, item in enumerate(order.items.all(), 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=item.product.name)
        ws.cell(row=row, column=3, value=item.product.article)
        ws.cell(row=row, column=4, value=float(item.quantity))
        ws.cell(row=row, column=5, value=item.product.unit)
        ws.cell(row=row, column=6, value=float(item.price))
        ws.cell(row=row, column=7, value=float(item.total_price))
        row += 1
    
    # Итого
    total_row = row + 1
    ws.cell(row=total_row, column=6, value='ИТОГО:')
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=float(order.total_amount))
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=7, max_row=total_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            if cell.row > 7 and cell.row < total_row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


