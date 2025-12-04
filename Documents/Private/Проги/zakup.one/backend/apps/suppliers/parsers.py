import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from typing import List, Dict, Optional, Any
import logging
import re
import math
import hashlib

logger = logging.getLogger(__name__)

# Цвет фона для категорий (золотисто-желтый и похожие оттенки)
# Формат: RGB в формате AARRGGBB (с префиксом FF для альфа-канала)
CATEGORY_BACKGROUND_COLORS = [
    'FFFFD700',  # Gold
    'FFFFFF00',  # Yellow
    'FFFFE4B5',  # Moccasin
    'FFFFDAB9',  # PeachPuff
    'FFFFE4E1',  # MistyRose
    'FFFFF8DC',  # Cornsilk
    'FFFFFACD',  # LemonChiffon
    'FFEEE8AA',  # PaleGoldenrod
    'FFDAA520',  # Goldenrod
    'FFB8860B',  # DarkGoldenrod
    'FFFFEBCD',  # BlanchedAlmond
    'FFFFEFD5',  # PapayaWhip
    'FFFFF5EE',  # Seashell
    'FFFFF8F0',  # FloralWhite
    'FFFFFFF0',  # Ivory
]

def is_category_color(fill) -> bool:
    """Проверяет, является ли цвет заливки цветом категории (золотисто-желтый или серый)"""
    if not fill or fill.patternType != 'solid':
        return False
    
    if not fill.start_color:
        return False
    
    # Проверяем RGB цвет
    if hasattr(fill.start_color, 'rgb') and fill.start_color.rgb:
        rgb = str(fill.start_color.rgb).upper()
        
        # Убираем префикс FF если есть (формат AARRGGBB)
        rgb_clean = rgb[2:] if rgb.startswith('FF') and len(rgb) == 8 else rgb
        
        # Проверяем точное совпадение
        if rgb in CATEGORY_BACKGROUND_COLORS:
            return True
        
        # Проверяем без префикса альфа-канала
        if rgb_clean in [c[2:] if len(c) == 8 else c for c in CATEGORY_BACKGROUND_COLORS]:
            return True
        
        # Проверяем похожие оттенки (золотисто-желтые)
        # Извлекаем RGB компоненты
        if len(rgb_clean) == 6:
            try:
                r = int(rgb_clean[0:2], 16)
                g = int(rgb_clean[2:4], 16)
                b = int(rgb_clean[4:6], 16)
                
                # Золотисто-желтый: высокий R и G, низкий B
                # Пример: FFC000 (RGB: 255, 192, 0) - золотисто-желтый
                if r >= 200 and g >= 150 and b <= 100:
                    # Дополнительная проверка: G должен быть близок к R
                    if abs(r - g) < 100:
                        return True
                
                # Серый цвет для категорий: все компоненты примерно равны и в диапазоне 150-220
                # Пример: D3D3D3 (RGB: 211, 211, 211) - светло-серый
                if 150 <= r <= 220 and 150 <= g <= 220 and 150 <= b <= 220:
                    # Разница между компонентами не должна быть большой (серый цвет)
                    if abs(r - g) < 30 and abs(g - b) < 30 and abs(r - b) < 30:
                        return True
            except ValueError:
                pass
    
    # Проверяем индекс цвета (для стандартных цветов Excel)
    if hasattr(fill.start_color, 'index') and fill.start_color.index is not None:
        color_index = fill.start_color.index
        # Индексы для желтых/золотых оттенков в Excel (расширенный диапазон)
        yellow_indices = list(range(44, 54))  # 44-53
        # Индексы для серых оттенков
        gray_indices = list(range(22, 26))  # 22-25 обычно серые
        if color_index in yellow_indices or color_index in gray_indices:
            return True
    
    # Проверяем theme color (для тем Excel)
    if hasattr(fill.start_color, 'theme') and fill.start_color.theme is not None:
        # Theme 4 обычно соответствует желтым оттенкам
        if fill.start_color.theme in [4, 5, 6]:
            return True
    
    return False


def is_white_background(fill) -> bool:
    """Проверяет, является ли фон ячейки белым (или отсутствует заливка)"""
    if not fill or fill.patternType != 'solid':
        return True  # Нет заливки = белый фон
    
    if not fill.start_color:
        return True  # Нет цвета = белый фон
    
    # Проверяем RGB цвет
    if hasattr(fill.start_color, 'rgb') and fill.start_color.rgb:
        rgb = str(fill.start_color.rgb).upper()
        rgb_clean = rgb[2:] if rgb.startswith('FF') and len(rgb) == 8 else rgb
        
        if len(rgb_clean) == 6:
            try:
                r = int(rgb_clean[0:2], 16)
                g = int(rgb_clean[2:4], 16)
                b = int(rgb_clean[4:6], 16)
                
                # Белый цвет: все компоненты близки к 255 (>= 250)
                if r >= 250 and g >= 250 and b >= 250:
                    return True
            except ValueError:
                pass
    
    # Проверяем индекс цвета
    if hasattr(fill.start_color, 'index') and fill.start_color.index is not None:
        # Индекс 0 или None обычно означает белый/нет заливки
        if fill.start_color.index == 0:
            return True
    
    return False


def is_gray_text(font) -> bool:
    """Проверяет, является ли цвет шрифта серым"""
    if not font or not font.color:
        return False  # Нет цвета = черный по умолчанию
    
    # Проверяем RGB цвет
    if hasattr(font.color, 'rgb') and font.color.rgb:
        rgb = str(font.color.rgb).upper()
        rgb_clean = rgb[2:] if rgb.startswith('FF') and len(rgb) == 8 else rgb
        
        if len(rgb_clean) == 6:
            try:
                r = int(rgb_clean[0:2], 16)
                g = int(rgb_clean[2:4], 16)
                b = int(rgb_clean[4:6], 16)
                
                # Серый цвет шрифта: все компоненты примерно равны и в диапазоне 100-200
                # Черный цвет: все компоненты < 50
                # Серый: компоненты в диапазоне 100-200 и примерно равны
                if 100 <= r <= 200 and 100 <= g <= 200 and 100 <= b <= 200:
                    if abs(r - g) < 30 and abs(g - b) < 30 and abs(r - b) < 30:
                        return True
            except ValueError:
                pass
    
    return False


def is_black_text(font) -> bool:
    """Проверяет, является ли цвет шрифта черным (или отсутствует)"""
    if not font or not font.color:
        return True  # Нет цвета = черный по умолчанию
    
    # Проверяем RGB цвет
    if hasattr(font.color, 'rgb') and font.color.rgb:
        rgb = str(font.color.rgb).upper()
        rgb_clean = rgb[2:] if rgb.startswith('FF') and len(rgb) == 8 else rgb
        
        if len(rgb_clean) == 6:
            try:
                r = int(rgb_clean[0:2], 16)
                g = int(rgb_clean[2:4], 16)
                b = int(rgb_clean[4:6], 16)
                
                # Черный цвет: все компоненты < 50
                if r < 50 and g < 50 and b < 50:
                    return True
            except ValueError:
                pass
    
    # Если индекс цвета 1 (черный) или None (по умолчанию черный)
    if hasattr(font.color, 'index'):
        if font.color.index is None or font.color.index == 1:
            return True
    
    return False


class ExcelPriceListParser:
    """Парсер для Excel прайс-листов
    
    Правила определения категорий:
    - Если ячейка имеет цвет фона из списка CATEGORY_BACKGROUND_COLORS - это категория/подкатегория
    - Если в колонке 0 есть название, но в колонке 2 нет единицы измерения - это категория
    - Если в колонке 0 есть название И в колонке 2 есть единица измерения - это товар
    """
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.products = []
        self.categories = []
        self.current_category = None
        self.workbook = None
        self.worksheet = None
        
    def parse(self, sheet_name: str = None) -> Dict:
        """
        Парсит Excel файл и возвращает словарь с товарами и категориями
        
        Новый алгоритм:
        1. Ищет строку с заголовком "Товар | Ед.изм | Цена"
        2. Берет все строки до последней строки с ценой
        3. Определяет категории: строки без цены, но с названием (или с цветом фона категории)
        4. Определяет товары: строки с названием и ценой
        
        Возвращает:
        {
            'products': [{'name': str, 'article': str, 'unit': str, 'price': float, 'category': str}],
            'categories': [str],
            'total_products': int
        }
        """
        try:
            # Открываем Excel файл для чтения цветов
            logger.info(f"Чтение файла: {self.file_path}")
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            if sheet_name:
                self.worksheet = self.workbook[sheet_name]
            else:
                self.worksheet = self.workbook.active
            
            # Читаем Excel файл через pandas для данных
            excel_data = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            
            # Проверяем, что получили DataFrame, а не словарь (если несколько листов)
            if isinstance(excel_data, dict):
                # Если словарь, берем первый лист или указанный лист
                if sheet_name and sheet_name in excel_data:
                    df = excel_data[sheet_name]
                else:
                    # Берем первый лист
                    first_sheet = list(excel_data.keys())[0]
                    df = excel_data[first_sheet]
                    logger.info(f"Файл содержит несколько листов. Используется лист: {first_sheet}")
            else:
                df = excel_data
            
            # Убеждаемся, что df - это DataFrame
            if not isinstance(df, pd.DataFrame):
                raise ValueError(f"Ожидался DataFrame, получен {type(df)}")
            
            logger.info(f"Файл прочитан. Строк: {len(df)}, Колонок: {len(df.columns) if len(df) > 0 else 0}")
            
            # Ищем строку с заголовком "Товар"
            header_row = self._find_header_row_new(df)
            
            if header_row is None:
                logger.warning("Не найдена строка с заголовком 'Товар'. Пробуем старый метод.")
                header_row = self._find_header_row(df)
                if header_row is None:
                    header_row = 8  # Fallback: строка 9 (0-based индекс 8)
            
            logger.info(f"Найдена строка заголовков: {header_row + 1} (индекс {header_row})")
            
            # Получаем индексы колонок
            header = df.iloc[header_row]
            try:
                product_col = header[header == "Товар"].index[0]
                unit_col = header[header == "Ед.изм"].index[0]
                price_col = header[header == "Цена"].index[0]
            except (IndexError, KeyError):
                # Если не нашли точные заголовки, пробуем найти по ключевым словам
                logger.warning("Не найдены точные заголовки. Пробуем найти по ключевым словам.")
                product_col, unit_col, price_col = self._find_columns_by_keywords(df, header_row)
            
            logger.info(f"Колонки: Товар={product_col}, Ед.изм={unit_col}, Цена={price_col}")
            
            # Находим последнюю строку с ценой
            price_series = df.iloc[:, price_col]
            last_price_idx = price_series.last_valid_index()
            if last_price_idx is None:
                logger.warning("Не найдена ни одна строка с ценой. Используем все строки после заголовка.")
                last_price_idx = len(df) - 1
            
            logger.info(f"Последняя строка с ценой: {last_price_idx + 1} (индекс {last_price_idx})")
            
            # Парсим данные
            self._parse_data_new(df, header_row + 1, last_price_idx + 1, product_col, unit_col, price_col)
            
            logger.info(f"Парсинг завершен. Товаров: {len(self.products)}, категорий: {len(self.categories)}")
            
            # Закрываем workbook
            if self.workbook:
                self.workbook.close()
            
            return {
                'products': self.products,
                'categories': self.categories,
                'total_products': len(self.products)
            }
            
        except Exception as e:
            logger.error(f"Ошибка парсинга Excel файла: {str(e)}", exc_info=True)
            if self.workbook:
                self.workbook.close()
            raise
    
    def _find_header_row_new(self, df: pd.DataFrame) -> Optional[int]:
        """Находит строку с заголовком 'Товар' (новый метод)"""
        for idx, row in df.iterrows():
            # Ищем строку, где в любой ячейке есть слово "Товар"
            if (df.iloc[idx] == "Товар").any():
                return idx
        return None
    
    def _find_header_row(self, df: pd.DataFrame) -> Optional[int]:
        """Находит строку с заголовками таблицы (старый метод)"""
        header_keywords = ['товар', 'ед.изм', 'ед', 'цена', 'единица', 'измерения']
        
        for idx, row in df.iterrows():
            row_text = ' '.join([str(cell).lower() if pd.notna(cell) else '' for cell in row.values])
            
            # Проверяем, содержит ли строка ключевые слова заголовков
            if any(keyword in row_text for keyword in header_keywords):
                # Проверяем, что это действительно заголовки (не слишком много пустых ячеек)
                non_empty = sum(1 for cell in row.values if pd.notna(cell) and str(cell).strip())
                if non_empty >= 2:  # Минимум 2 колонки должны быть заполнены
                    return idx
        
        return None
    
    def _find_columns_by_keywords(self, df: pd.DataFrame, header_row: int) -> tuple:
        """Находит колонки по ключевым словам"""
        header = df.iloc[header_row]
        product_col = None
        unit_col = None
        price_col = None
        
        for idx, val in enumerate(header):
            val_str = str(val).lower().strip() if pd.notna(val) else ''
            if 'товар' in val_str and product_col is None:
                product_col = idx
            elif ('ед' in val_str or 'единиц' in val_str) and unit_col is None:
                unit_col = idx
            elif 'цена' in val_str and price_col is None:
                price_col = idx
        
        # Если не нашли, используем стандартные индексы
        if product_col is None:
            product_col = 0
        if unit_col is None:
            unit_col = 2
        if price_col is None:
            price_col = 3
        
        return product_col, unit_col, price_col
    
    def _to_str_or_none(self, value) -> Optional[str]:
        """Нормализуем строку: если NaN — вернём None, иначе обрежем пробелы."""
        if isinstance(value, float) and math.isnan(value):
            return None
        if value is None:
            return None
        s = str(value).strip()
        return s or None
    
    def _to_float_or_nan(self, value) -> float:
        """Пытаемся привести значение к float (с учётом пробелов и запятых)."""
        if isinstance(value, float) or isinstance(value, int):
            return float(value)
        
        if value is None:
            return float("nan")
        
        s = str(value).strip()
        if not s:
            return float("nan")
        
        s = s.replace(" ", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return float("nan")
    
    def _parse_data_new(self, df: pd.DataFrame, start_row: int, end_row: int, 
                       product_col: int, unit_col: int, price_col: int):
        """
        Улучшенный метод парсинга данных с учетом цветов ячеек
        
        Правила для формата Стройдвор:
        1. Если цвет ячейки не белый - это не товар (пропускаем)
        2. Если цвет ячейки серый И цвет шрифта серый - это категория/подкатегория
        3. Если цвет ячейки белый И цвет шрифта черный:
           - Данные в столбцах 1-2 (индексы 0-1) - название товара
           - На следующей строке - единица измерения
           - На следующей строке - цена
        """
        current_category: Optional[str] = None
        current_subcategory: Optional[str] = None
        
        # Берём только строки с данными (между заголовком и последней ценой)
        data = df.iloc[start_row:end_row].copy()
        
        i = 0
        data_list = list(data.iterrows())
        
        while i < len(data_list):
            idx, row = data_list[i]
            
            # row - это pandas Series, используем iloc для доступа по индексу
            # Получаем значения из столбцов 1-2 (индексы 0-1)
            col0_value = None
            col1_value = None
            if len(row) > 0:
                try:
                    col0_value = self._to_str_or_none(row.iloc[0])  # Столбец 1
                except (IndexError, KeyError):
                    pass
            if len(row) > 1:
                try:
                    col1_value = self._to_str_or_none(row.iloc[1])  # Столбец 2
                except (IndexError, KeyError):
                    pass
            
            # Получаем единицу измерения и цену из текущей строки (на случай стандартного формата)
            raw_unit = None
            raw_price = None
            try:
                if unit_col < len(row):
                    raw_unit = row.iloc[unit_col] if hasattr(row, 'iloc') else row[unit_col]
            except (IndexError, KeyError):
                pass
            try:
                if price_col < len(row):
                    raw_price = row.iloc[price_col] if hasattr(row, 'iloc') else row[price_col]
            except (IndexError, KeyError):
                pass
            
            unit = self._to_str_or_none(raw_unit)
            price = self._to_float_or_nan(raw_price)
            
            # Получаем информацию о цветах ячеек
            is_white_bg = True
            is_category_bg = False
            has_gray_text = False
            has_black_text = True
            
            if self.worksheet:
                try:
                    # Проверяем ячейку в столбце 1 (индекс 0)
                    cell_col0 = self.worksheet.cell(row=idx+1, column=1)
                    fill_col0 = cell_col0.fill
                    font_col0 = cell_col0.font
                    
                    is_white_bg = is_white_background(fill_col0)
                    is_category_bg = is_category_color(fill_col0)
                    has_gray_text = is_gray_text(font_col0)
                    has_black_text = is_black_text(font_col0)
                except Exception as e:
                    logger.debug(f"Ошибка чтения цвета ячейки row={idx+1}, col=1: {str(e)}")
            
            # Пропускаем полностью пустые строки
            if not col0_value and not col1_value and not unit and math.isnan(price):
                i += 1
                continue
            
            # ПРАВИЛО 1: Если цвет ячейки не белый - это не товар (пропускаем или категория)
            if not is_white_bg:
                # Если это категория по цвету (серый фон + серый шрифт)
                if is_category_bg and has_gray_text and col0_value:
                    current_category = col0_value
                    if current_category not in self.categories:
                        self.categories.append(current_category)
                    logger.debug(f"Найдена категория по цвету (серый фон + серый шрифт): '{current_category}'")
                    current_subcategory = None  # Сбрасываем подкатегорию при новой категории
                i += 1
                continue
            
            # ПРАВИЛО 2: Если белый фон + черный шрифт + есть данные в столбцах 1-2
            if is_white_bg and has_black_text and (col0_value or col1_value):
                product_name = col0_value or col1_value
                product_unit = None
                product_price = None
                skip_next_rows = 0
                
                # ВАРИАНТ 1: Проверяем текущую строку (стандартный формат)
                if unit and self._is_unit_measurement(unit):
                    product_unit = unit
                if not math.isnan(price):
                    product_price = price
                
                # ВАРИАНТ 2: Если не нашли в текущей строке, проверяем следующие строки
                if not product_unit or math.isnan(product_price) if product_price is None else False:
                    # Проверяем следующую строку на единицу измерения
                    if i + 1 < len(data_list):
                        next_idx, next_row = data_list[i + 1]
                        
                        # Проверяем цвет следующей строки (должна быть белой для товара)
                        next_is_white = True
                        if self.worksheet:
                            try:
                                next_cell = self.worksheet.cell(row=next_idx+1, column=1)
                                next_is_white = is_white_background(next_cell.fill)
                            except:
                                pass
                        
                        if next_is_white:
                            # Проверяем столбец единицы измерения
                            next_unit = None
                            try:
                                if unit_col < len(next_row):
                                    next_unit = self._to_str_or_none(next_row.iloc[unit_col] if hasattr(next_row, 'iloc') else next_row[unit_col])
                            except (IndexError, KeyError):
                                pass
                            if next_unit and self._is_unit_measurement(next_unit):
                                product_unit = next_unit
                                skip_next_rows = 1
                                
                                # Проверяем следующую строку на цену
                                if i + 2 < len(data_list):
                                    price_idx, price_row = data_list[i + 2]
                                    
                                    # Проверяем цвет строки с ценой
                                    price_is_white = True
                                    if self.worksheet:
                                        try:
                                            price_cell = self.worksheet.cell(row=price_idx+1, column=1)
                                            price_is_white = is_white_background(price_cell.fill)
                                        except:
                                            pass
                                    
                                    if price_is_white:
                                        try:
                                            if price_col < len(price_row):
                                                price_val = price_row.iloc[price_col] if hasattr(price_row, 'iloc') else price_row[price_col]
                                                product_price = self._to_float_or_nan(price_val)
                                            else:
                                                product_price = float("nan")
                                        except (IndexError, KeyError):
                                            product_price = float("nan")
                                        if not math.isnan(product_price):
                                            skip_next_rows = 2
                
                # Если нашли товар (есть название, единица измерения и цена)
                if product_name and product_unit and product_price and not math.isnan(product_price):
                    # Нормализуем единицу измерения
                    unit_normalized = self._normalize_unit(product_unit)
                    
                    # Генерируем артикул
                    article = self._generate_article(product_name)
                    
                    # Используем категорию или подкатегорию
                    category = current_subcategory or current_category
                    
                    product = {
                        'name': product_name,
                        'article': article,
                        'unit': unit_normalized,
                        'price': float(product_price),
                        'category': category
                    }
                    
                    self.products.append(product)
                    logger.debug(f"Извлечен товар: {product_name}, цена: {product_price}, единица: {unit_normalized}, категория: {category}")
                    
                    # Пропускаем строки с единицей измерения и ценой, если они были на следующих строках
                    i += skip_next_rows
                else:
                    # Если есть название, но нет единицы измерения или цены - возможно подкатегория
                    if product_name and not product_unit:
                        current_subcategory = product_name
                        logger.debug(f"Найдена подкатегория: '{current_subcategory}'")
            
            i += 1
        
        logger.info(f"Распознано товаров: {len(self.products)}, категорий: {len(self.categories)}")
        if len(self.products) == 0:
            logger.warning("Не найдено ни одного товара! Проверьте структуру файла.")
    
    def _parse_data(self, df: pd.DataFrame, start_row: int):
        """Парсит данные товаров из DataFrame
        
        Логика для Стройдвор (реальная структура файла):
        - Колонка 0 (индекс 0): название товара/категории
        - Колонка 1 (индекс 1): обычно пустая
        - Колонка 2 (индекс 2): единица измерения (для товаров) или может быть число (для категорий)
        - Колонка 3 (индекс 3): цена
        
        Правила определения категорий (в порядке приоритета):
        1. Если ячейка в колонке 0 имеет цвет фона из CATEGORY_BACKGROUND_COLORS - это категория
        2. Если в колонке 0 есть запись, но в колонке 2 нет единицы измерения - это категория
        3. Если в колонке 0 есть запись И в колонке 2 есть единица измерения - это товар
        """
        current_category = None
        
        for idx in range(start_row, len(df)):
            row = df.iloc[idx]
            
            # Пропускаем полностью пустые строки
            if row.isna().all():
                continue
            
            # Получаем значения конкретных колонок (по индексу)
            # Колонка 0 (индекс 0) - название товара/категории
            # Колонка 2 (индекс 2) - единица измерения
            # Колонка 3 (индекс 3) - цена
            
            col0_value = str(row.iloc[0]).strip() if len(row) > 0 and pd.notna(row.iloc[0]) else ''
            col2_value = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ''
            col3_value = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else ''
            
            # Пропускаем строки, где нет названия
            if not col0_value:
                continue
            
            # ПРИОРИТЕТ 1: Проверяем цвет фона ячейки в колонке 0 (A)
            # Excel строки начинаются с 1, а pandas индексы с 0, поэтому idx+1
            is_category_by_color = False
            if self.worksheet:
                try:
                    cell = self.worksheet.cell(row=idx+1, column=1)  # Колонка A (1)
                    fill = cell.fill
                    is_category_by_color = is_category_color(fill)
                    if is_category_by_color:
                        color_info = 'unknown'
                        if hasattr(fill.start_color, 'rgb') and fill.start_color.rgb:
                            color_info = fill.start_color.rgb
                        elif hasattr(fill.start_color, 'index'):
                            color_info = f'index={fill.start_color.index}'
                        logger.debug(f"Найдена категория по цвету фона: '{col0_value}', цвет: {color_info}")
                except Exception as e:
                    logger.debug(f"Ошибка чтения цвета ячейки row={idx+1}, col=1: {str(e)}")
            
            # Если ячейка имеет цвет категории - это категория, пропускаем проверку единицы измерения
            if is_category_by_color:
                current_category = col0_value
                if current_category not in self.categories:
                    self.categories.append(current_category)
                logger.debug(f"Найдена категория (по цвету): {current_category}")
                continue
            
            # ПРИОРИТЕТ 2: Проверяем, является ли значение в колонке 2 единицей измерения
            is_unit = self._is_unit_measurement(col2_value)
            
            # Определяем тип строки по логике
            # Если есть название (колонка 0), но нет единицы измерения в колонке 2 - это категория
            if col0_value and not is_unit:
                # Это категория
                current_category = col0_value
                if current_category not in self.categories:
                    self.categories.append(current_category)
                logger.debug(f"Найдена категория (по отсутствию единицы): {current_category}")
                
            # Если есть название (колонка 0) И есть единица измерения в колонке 2 - это товар
            elif col0_value and is_unit:
                # Это товар
                product = self._extract_product_stroydvor(row, current_category)
                if product:
                    self.products.append(product)
        
        logger.info(f"Распознано товаров: {len(self.products)}, категорий: {len(self.categories)}")
        if len(self.products) == 0:
            logger.warning("Не найдено ни одного товара! Проверьте структуру файла.")
    
    def _is_unit_measurement(self, value: str) -> bool:
        """Проверяет, является ли значение единицей измерения"""
        if not value:
            return False
        
        value_upper = value.strip().upper()
        
        # Список единиц измерения
        units = ['ШТ', 'ШТУК', 'ШТУКИ', 'М', 'МЕТР', 'МЕТРЫ', 'М2', 'М²', 'М.КВ', 'М.КВ.', 'КВ.М', 'КВ.М.', 
                 'М3', 'М³', 'М.КУБ', 'М.КУБ.', 'КГ', 'КИЛОГРАММ', 'КИЛОГРАММЫ', 'Т', 'ТОННА', 'ТОННЫ',
                 'Л', 'ЛИТР', 'ЛИТРЫ', 'МЛ', 'МИЛЛИЛИТР']
        
        if value_upper in units:
            return True
        
        # Проверка по содержимому
        if 'шт' in value.lower():
            return True
        elif 'м²' in value or 'м2' in value.upper() or ('кв' in value.lower() and 'м' in value.lower()):
            return True
        elif 'м³' in value or 'м3' in value.upper() or 'куб' in value.lower():
            return True
        elif 'кг' in value.lower():
            return True
        elif 'т' in value.lower() and ('тонн' in value.lower() or len(value) <= 3):
            return True
        elif 'л' in value.lower() and 'мл' not in value.lower():
            return True
        elif 'мл' in value.lower():
            return True
        elif 'м' in value.lower() and len(value) <= 5:  # Короткие значения с "м" могут быть единицами
            return True
        
        # Если это число, это не единица измерения
        try:
            float(re.sub(r'[^\d.]', '', value))
            return False
        except:
            pass
        
        return False
    
    
    def _extract_product_stroydvor(self, row: pd.Series, category: Optional[str]) -> Optional[Dict]:
        """Извлекает данные товара из строки для формата Стройдвор
        
        Структура (реальная):
        - Колонка 0 (индекс 0): название товара
        - Колонка 2 (индекс 2): единица измерения
        - Колонка 3 (индекс 3): цена
        """
        try:
            # Получаем значения по индексам колонок
            product_name = str(row.iloc[0]).strip() if len(row) > 0 and pd.notna(row.iloc[0]) else ''
            unit_raw = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else ''
            price_raw = str(row.iloc[3]).strip() if len(row) > 3 and pd.notna(row.iloc[3]) else ''
            
            if not product_name:
                return None
            
            if not unit_raw or not self._is_unit_measurement(unit_raw):
                # Если нет единицы измерения, это не товар (должно быть обработано как категория)
                logger.debug(f"Пропущена строка без единицы измерения: {product_name}, col2={unit_raw}")
                return None
            
            # Парсим цену
            price = None
            try:
                # Убираем пробелы и пробуем преобразовать в число
                clean_price = re.sub(r'[\s]', '', price_raw)  # Убираем пробелы
                # Пробуем извлечь число (может быть с разделителями тысяч)
                numbers = re.findall(r'\d+[.,]?\d*', clean_price)
                if numbers:
                    # Берем первое число
                    num_str = numbers[0].replace(',', '.')
                    price = float(num_str)
            except Exception as e:
                logger.debug(f"Не удалось распарсить цену из '{price_raw}': {str(e)}")
            
            if price is None or price <= 0:
                logger.debug(f"Не найдена валидная цена в строке: название={product_name}, цена={price_raw}")
                return None
            
            # Нормализуем единицу измерения
            unit = self._normalize_unit(unit_raw)
            
            # Генерируем артикул из названия
            article = self._generate_article(product_name)
            
            product = {
                'name': product_name,
                'article': article,
                'unit': unit,
                'price': price,
                'category': category
            }
            
            logger.debug(f"Извлечен товар: {product_name}, цена: {price}, единица: {unit}, категория: {category}")
            return product
            
        except Exception as e:
            logger.error(f"Ошибка извлечения товара из строки: {str(e)}")
            return None
    
    def _normalize_unit(self, unit_raw: str) -> str:
        """Нормализует единицу измерения к стандартному формату"""
        unit_upper = unit_raw.strip().upper()
        
        # Прямые совпадения
        unit_map = {
            'ШТ': 'шт',
            'ШТУК': 'шт',
            'ШТУКИ': 'шт',
            'М': 'м',
            'МЕТР': 'м',
            'МЕТРЫ': 'м',
            'М2': 'м²',
            'М²': 'м²',
            'М.КВ': 'м²',
            'М.КВ.': 'м²',
            'КВ.М': 'м²',
            'КВ.М.': 'м²',
            'М3': 'м³',
            'М³': 'м³',
            'М.КУБ': 'м³',
            'М.КУБ.': 'м³',
            'КГ': 'кг',
            'КИЛОГРАММ': 'кг',
            'КИЛОГРАММЫ': 'кг',
            'Т': 'т',
            'ТОННА': 'т',
            'ТОННЫ': 'т',
            'Л': 'л',
            'ЛИТР': 'л',
            'ЛИТРЫ': 'л',
            'МЛ': 'мл',
            'МИЛЛИЛИТР': 'мл',
        }
        
        if unit_upper in unit_map:
            return unit_map[unit_upper]
        
        # Проверка по содержимому
        if 'шт' in unit_raw.lower():
            return 'шт'
        elif 'м²' in unit_raw or 'м2' in unit_raw.upper() or 'кв' in unit_raw.lower():
            return 'м²'
        elif 'м³' in unit_raw or 'м3' in unit_raw.upper() or 'куб' in unit_raw.lower():
            return 'м³'
        elif 'кг' in unit_raw.lower():
            return 'кг'
        elif 'т' in unit_raw.lower() and 'тонн' in unit_raw.lower():
            return 'т'
        elif 'л' in unit_raw.lower() and 'мл' not in unit_raw.lower():
            return 'л'
        elif 'мл' in unit_raw.lower():
            return 'мл'
        elif 'м' in unit_raw.lower():
            return 'м'
        
        # По умолчанию
        return unit_raw.lower()
    
    def _generate_article(self, name: str) -> str:
        """Генерирует артикул из названия товара"""
        # Берем первые буквы слов и добавляем хеш
        words = name.split()
        if words:
            # Берем первые 3-4 буквы из первых слов
            article_parts = []
            for word in words[:3]:
                if word and word[0].isalnum():
                    article_parts.append(word[:3].upper())
            
            if article_parts:
                article = ''.join(article_parts)
                # Добавляем короткий хеш для уникальности
                import hashlib
                hash_part = hashlib.md5(name.encode()).hexdigest()[:4].upper()
                return f"{article}-{hash_part}"
        
        # Если не получилось, используем хеш
        import hashlib
        return hashlib.md5(name.encode()).hexdigest()[:8].upper()

